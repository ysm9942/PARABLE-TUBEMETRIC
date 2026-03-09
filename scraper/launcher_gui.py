"""
PARABLE-TUBEMETRIC — 통합 GUI
Vercel 대시보드의 모든 기능을 로컬 exe에 구현:
  - 채널 통합 분석  (YouTube API)
  - 단일 영상 분석  (YouTube API)
  - 광고 영상 분석  (YouTube API + NLP)
  - 로컬 스크래퍼   (undetected-chromedriver → GitHub push)
  - 라이브 지표 분석 (CHZZK/SOOP · viewership.softc.one)
  - 데이터 대시보드 (결과 표시 + Excel 내보내기)
"""
import sys
import re
import time
import random
import pickle
import threading
import socket
import getpass
import platform
import json
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog

# ── PyInstaller frozen path ────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent

TARGETS_FILE = SCRIPT_DIR / "targets.txt"

# ── 색상 팔레트 ────────────────────────────────────────────────────────────────
BG      = "#0a0a0a"
BG2     = "#111111"
BG3     = "#1a1a1a"
BG4     = "#222222"
BG5     = "#141414"   # 짝수 행 배경 (스트라이프)
BORDER  = "#222222"
BORDER2 = "#2a2a2a"   # 카드 경계선
FG      = "#f4f4f5"
FG_DIM  = "#d4d4d8"
FG_MUTE = "#71717a"
ACCENT  = "#dc2626"
ACCENT2 = "#ef4444"
ACCENT_DIM = "#7f1d1d"  # 비활성 액센트
GREEN   = "#22c55e"
BLUE    = "#60a5fa"
YELLOW  = "#facc15"

PIN_CORRECT = "5350"

MACHINE_INFO = {
    "operator": getpass.getuser(),
    "hostname": socket.gethostname(),
    "os": f"{platform.system()} {platform.release()}",
}


# ── 인증 정보 로드 (config.py 내장값 우선) ────────────────────────────────────
def _load_credentials():
    try:
        from config import get_github_token, GITHUB_REPO
        token = get_github_token()
        if token:
            return token, GITHUB_REPO
    except ImportError:
        pass
    import os
    return os.environ.get("GITHUB_TOKEN", ""), os.environ.get("GITHUB_REPO", "")


def _load_yt_api_key() -> str:
    try:
        from config import YOUTUBE_API_KEY
        if YOUTUBE_API_KEY:
            return YOUTUBE_API_KEY
    except (ImportError, AttributeError):
        pass
    import os
    return os.environ.get("YOUTUBE_API_KEY", "")


def _load_gsheet_url() -> str:
    # SCRIPT_DIR을 sys.path에 추가해 실행 위치에 무관하게 config를 찾음
    import importlib, sys as _sys
    _dir = str(SCRIPT_DIR)
    _added = _dir not in _sys.path
    if _added:
        _sys.path.insert(0, _dir)
    try:
        cfg = importlib.import_module("config")
        url = getattr(cfg, "GSHEET_URL", "")
        if url:
            return url
    except Exception:
        pass
    finally:
        if _added and _dir in _sys.path:
            _sys.path.remove(_dir)
    return ""


# ── 숫자 포맷 ─────────────────────────────────────────────────────────────────
def fmt_num(n) -> str:
    try:
        n = int(n)
    except (TypeError, ValueError):
        return "0"
    if n >= 1_000_000:
        return f"{n/1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n/1_000:.1f}K"
    return f"{n:,}"


# ── ConsoleWindow ─────────────────────────────────────────────────────────────
class ConsoleWindow(tk.Toplevel):
    def __init__(self, master, title="실행 중"):
        super().__init__(master)
        self.title(f"PARABLE-TUBEMETRIC  |  {title}")
        self.geometry("800x500")
        self.configure(bg="#0c0c0c")
        self.resizable(True, True)
        self._closed = False
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        bar = tk.Frame(self, bg="#161616")
        bar.pack(fill="x")
        tk.Label(bar, text=f"  ■  {title}", font=("Consolas", 9, "bold"),
                 bg="#161616", fg="#555", pady=7, anchor="w").pack(fill="x", padx=4)

        self.log = scrolledtext.ScrolledText(
            self, font=("Consolas", 10),
            bg="#0c0c0c", fg="#cccccc",
            insertbackground="white", relief="flat", padx=14, pady=10,
            state="disabled",
        )
        self.log.pack(fill="both", expand=True)
        self.log.tag_configure("ok",   foreground="#4ade80")
        self.log.tag_configure("err",  foreground="#f87171")
        self.log.tag_configure("info", foreground="#60a5fa")
        self.log.tag_configure("dim",  foreground="#555555")

    def append(self, text: str, tag: str = ""):
        def _do():
            self.log.configure(state="normal")
            if not tag:
                low = text.lower()
                if any(k in low for k in ["완료", "✓", "done", "push"]):
                    t = "ok"
                elif any(k in low for k in ["오류", "error", "fail", "✗"]):
                    t = "err"
                elif text.startswith("["):
                    t = "info"
                else:
                    t = "dim"
            else:
                t = tag
            self.log.insert("end", text, t)
            self.log.see("end")
            self.log.configure(state="disabled")
        self.after(0, _do)

    def finish(self, success: bool = True):
        msg = "\n✓  완료  —  5초 후 창이 닫힙니다.\n" if success else "\n✗  오류 발생  —  5초 후 창이 닫힙니다.\n"
        self.append(msg)
        if not self._closed:
            self.after(5000, self._safe_destroy)

    def _safe_destroy(self):
        if not self._closed:
            self.destroy()

    def _on_close(self):
        self._closed = True
        self.destroy()


# ── PIN 화면 ──────────────────────────────────────────────────────────────────
class PinScreen(tk.Frame):
    def __init__(self, master, on_success):
        super().__init__(master, bg=BG)
        self.on_success = on_success
        self._build()

    def _build(self):
        center = tk.Frame(self, bg=BG)
        center.place(relx=0.5, rely=0.5, anchor="center")

        icon = tk.Frame(center, bg=ACCENT, width=78, height=78)
        icon.pack_propagate(False)
        icon.pack(pady=(0, 18))
        tk.Label(icon, text="🔒", font=("Arial", 30), bg=ACCENT, fg="white").pack(expand=True)

        tk.Label(center, text="PARABLE-", font=("Arial", 26, "bold"), bg=BG, fg=FG).pack()
        tk.Label(center, text="TUBEMETRIC", font=("Arial", 26, "bold"), bg=BG, fg=ACCENT).pack()
        tk.Label(center, text="SYSTEM  LOCKED", font=("Arial", 8, "bold"),
                 bg=BG, fg=FG_DIM, pady=6).pack()
        tk.Frame(center, height=16, bg=BG).pack()

        border = tk.Frame(center, bg=BORDER, padx=1, pady=1)
        border.pack(fill="x", pady=(0, 10))
        self._pin = tk.StringVar()
        self._entry = tk.Entry(
            border, textvariable=self._pin, show="●",
            font=("Consolas", 24, "bold"), bg=BG3, fg=FG,
            insertbackground=ACCENT, relief="flat", justify="center", width=11,
        )
        self._entry.pack(ipady=14, ipadx=20)
        self._entry.bind("<Return>", lambda _: self._submit())
        self._entry.focus()

        tk.Button(center, text="UNLOCK  →", command=self._submit,
                  bg=ACCENT, fg="white", activebackground=ACCENT2,
                  font=("Arial", 11, "bold"), relief="flat",
                  cursor="hand2", pady=12).pack(fill="x")
        tk.Label(center, text="AUTHORIZED ACCESS ONLY",
                 font=("Arial", 7, "bold"), bg=BG, fg="#27272a", pady=14).pack()

    def _submit(self):
        if self._pin.get() == PIN_CORRECT:
            self.on_success()
        else:
            self._entry.configure(bg="#2d0a0a")
            self.after(500, lambda: self._entry.configure(bg=BG3))
            self._pin.set("")


# ── 공통 UI 헬퍼 ──────────────────────────────────────────────────────────────
def _section_header(parent, title: str, subtitle: str = ""):
    strip = tk.Frame(parent, bg=BG2)
    strip.pack(fill="x")
    inner = tk.Frame(strip, bg=BG2, padx=28, pady=16)
    inner.pack(fill="x")
    tk.Frame(inner, bg=ACCENT, width=3).pack(side="left", fill="y", padx=(0, 14))
    text_f = tk.Frame(inner, bg=BG2)
    text_f.pack(side="left", fill="x", expand=True)
    tk.Label(text_f, text=title, font=("Arial", 15, "bold"),
             bg=BG2, fg=FG, anchor="w").pack(fill="x")
    if subtitle:
        tk.Label(text_f, text=subtitle, font=("Arial", 9),
                 bg=BG2, fg=FG_DIM, anchor="w").pack(fill="x", pady=(3, 0))
    tk.Frame(parent, bg=BORDER, height=1).pack(fill="x")
    return strip


def _btn(parent, text, command, bg=BG3, fg=FG_DIM, bold=False, **kw):
    font = ("Arial", 10, "bold") if bold else ("Arial", 10)
    return tk.Button(parent, text=text, command=command,
                     bg=bg, fg=fg, activebackground=BG4,
                     activeforeground=FG, font=font, relief="flat",
                     cursor="hand2", **kw)


def _label(parent, text, font_size=9, color=FG_DIM, **kw):
    return tk.Label(parent, text=text,
                    font=("Arial", font_size), bg=BG, fg=color, **kw)


def _card(parent, **kw):
    f = tk.Frame(parent, bg=BG3, **kw)
    return f


# ── Treeview 스타일 ────────────────────────────────────────────────────────────
def _apply_treeview_style():
    # 한글 지원 폰트 선택 (Windows → Malgun Gothic, macOS → Apple SD Gothic Neo)
    import tkinter.font as _tkfont
    _available = _tkfont.families()
    _cell_font = ("Malgun Gothic", 9) if "Malgun Gothic" in _available else \
                 ("Apple SD Gothic Neo", 9) if "Apple SD Gothic Neo" in _available else \
                 ("TkDefaultFont", 9)
    _head_font = (_cell_font[0], 9, "bold")

    style = ttk.Style()
    style.theme_use("default")
    style.configure("Dark.Treeview",
        background=BG3, foreground=FG, fieldbackground=BG3,
        rowheight=30, font=_cell_font,
        borderwidth=0, relief="flat",
    )
    style.configure("Dark.Treeview.Heading",
        background=BG4, foreground=FG_DIM,
        font=_head_font, relief="flat",
        borderwidth=0, padding=(8, 6),
    )
    style.map("Dark.Treeview",
        background=[("selected", ACCENT_DIM)],
        foreground=[("selected", FG)],
    )
    style.map("Dark.Treeview.Heading",
        background=[("active", BG3)],
        foreground=[("active", FG)],
    )


# ── 채널 분석 탭 ──────────────────────────────────────────────────────────────
class ChannelTab(tk.Frame):
    def __init__(self, master, app):
        super().__init__(master, bg=BG)
        self.app = app
        self._build()

    def _build(self):
        _section_header(self, "채널 통합 분석",
                         "YouTube API → 쇼츠/롱폼 평균 조회수 · 영상 목록 수집")

        wrap = tk.Frame(self, bg=BG, padx=28)
        wrap.pack(fill="both", expand=True)

        # 채널 입력
        tk.Label(wrap, text="채널 목록  (한 줄에 하나 · @핸들, URL, UCxxx 모두 가능)",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(16, 4))
        border = tk.Frame(wrap, bg=ACCENT, padx=1, pady=1)
        border.pack(fill="x")
        self.ch_txt = tk.Text(border, height=5, font=("Consolas", 10),
                              bg=BG3, fg=FG, insertbackground=ACCENT,
                              relief="flat", padx=10, pady=8)
        self.ch_txt.pack(fill="both")

        # 옵션
        opt = tk.Frame(wrap, bg=BG, pady=10)
        opt.pack(fill="x")

        # 기간 필터
        tk.Label(opt, text="기간", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG, anchor="w").grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.period = tk.StringVar(value="all")
        for i, (val, lbl) in enumerate([("all","전체"),("90d","90일"),("30d","30일"),("7d","7일")]):
            rb = tk.Radiobutton(opt, text=lbl, variable=self.period, value=val,
                                bg=BG, fg=FG_DIM, activebackground=BG,
                                selectcolor=BG3, font=("Arial", 9),
                                activeforeground=FG)
            rb.grid(row=0, column=i+1, sticky="w", padx=4)

        self.use_date_filter = tk.BooleanVar(value=False)
        tk.Checkbutton(opt, text="날짜 필터 적용", variable=self.use_date_filter,
                       bg=BG, fg=FG_DIM, activebackground=BG, selectcolor=BG3,
                       font=("Arial", 9), activeforeground=FG
                       ).grid(row=0, column=6, sticky="w", padx=(20, 0))

        # 수집 개수
        tk.Label(opt, text="쇼츠", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG, anchor="w").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(8,0))
        self.shorts_count = tk.StringVar(value="30")
        tk.Entry(opt, textvariable=self.shorts_count, width=6,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).grid(row=1, column=1, sticky="w", pady=(8,0))

        tk.Label(opt, text="롱폼", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG, anchor="w").grid(row=1, column=2, sticky="w", padx=(16,8), pady=(8,0))
        self.longs_count = tk.StringVar(value="10")
        tk.Entry(opt, textvariable=self.longs_count, width=6,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).grid(row=1, column=3, sticky="w", pady=(8,0))

        self.use_count_filter = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text="개수 제한 사용", variable=self.use_count_filter,
                       bg=BG, fg=FG_DIM, activebackground=BG, selectcolor=BG3,
                       font=("Arial", 9), activeforeground=FG
                       ).grid(row=1, column=4, sticky="w", padx=(16, 0), pady=(8,0))

        self.use_shorts = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text="쇼츠 수집", variable=self.use_shorts,
                       bg=BG, fg=FG_DIM, activebackground=BG, selectcolor=BG3,
                       font=("Arial", 9), activeforeground=FG
                       ).grid(row=1, column=5, sticky="w", padx=(16, 0), pady=(8,0))

        self.use_longs = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text="롱폼 수집", variable=self.use_longs,
                       bg=BG, fg=FG_DIM, activebackground=BG, selectcolor=BG3,
                       font=("Arial", 9), activeforeground=FG
                       ).grid(row=1, column=6, sticky="w", padx=(16, 0), pady=(8,0))

        # 버튼
        btn_row = tk.Frame(wrap, bg=BG, pady=8)
        btn_row.pack(fill="x")
        self.run_btn = _btn(btn_row, "▶  분석 시작", self._run,
                            bg=ACCENT, fg="white", bold=True, padx=20, pady=10)
        self.run_btn.pack(side="left")

        self._status = tk.StringVar(value="대기 중")
        tk.Label(btn_row, textvariable=self._status, font=("Consolas", 9),
                 bg=BG, fg=FG_DIM).pack(side="left", padx=14)

        # API 키 상태
        self._api_status = tk.StringVar()
        tk.Label(wrap, textvariable=self._api_status,
                 font=("Arial", 8), bg=BG, fg=FG_DIM, anchor="w").pack(fill="x", pady=(4, 0))
        self._refresh_api_status()

    def _refresh_api_status(self):
        key = _load_yt_api_key()
        if key:
            self._api_status.set(f"✓  YouTube API Key 확인됨  ({key[:8]}...)")
        else:
            self._api_status.set("✗  YouTube API Key 없음  —  config.py 에 YOUTUBE_API_KEY 를 설정하고 재빌드하세요")

    def _run(self):
        channels = [
            ln.strip() for ln in self.ch_txt.get("1.0", "end").splitlines()
            if ln.strip() and not ln.strip().startswith("#")
        ]
        if not channels:
            messagebox.showwarning("입력 필요", "채널을 입력하세요.")
            return
        if not _load_yt_api_key():
            messagebox.showerror("API Key 없음", "YouTube API Key가 설정되지 않았습니다.")
            return

        self.run_btn.configure(state="disabled")
        self._status.set("분석 중...")

        shorts_cfg = {
            "enabled": self.use_shorts.get(),
            "target": int(self.shorts_count.get() or 30),
            "period": self.period.get(),
            "useDateFilter": self.use_date_filter.get(),
            "useCountFilter": self.use_count_filter.get(),
        }
        longs_cfg = {
            "enabled": self.use_longs.get(),
            "target": int(self.longs_count.get() or 10),
            "period": self.period.get(),
            "useDateFilter": self.use_date_filter.get(),
            "useCountFilter": self.use_count_filter.get(),
        }

        win = ConsoleWindow(self.winfo_toplevel(), "채널 통합 분석")
        win.append(f"[시작] 채널 {len(channels)}개 분석\n", "info")

        threading.Thread(
            target=self._analyze,
            args=(channels, shorts_cfg, longs_cfg, win),
            daemon=True,
        ).start()

    def _analyze(self, channels, shorts_cfg, longs_cfg, win):
        from youtube_api import get_channel_info, fetch_channel_stats
        results = []
        for ch in channels:
            win.append(f"\n[채널] {ch}\n", "info")
            try:
                info = get_channel_info(ch)
                win.append(f"  → {info['title']} (구독자 {fmt_num(info['subscriberCount'])})\n")
                stats = fetch_channel_stats(
                    info["uploadsPlaylistId"], shorts_cfg, longs_cfg,
                    progress_cb=lambda m: win.append(m + "\n"),
                )
                result = {**info, **stats, "status": "completed"}
                results.append(result)
                win.append(
                    f"  ✓ 쇼츠 {stats['shortsCount']}개 ({fmt_num(stats['avgShortsViews'])} avg) · "
                    f"롱폼 {stats['longCount']}개 ({fmt_num(stats['avgLongViews'])} avg)\n", "ok"
                )
            except Exception as e:
                win.append(f"  ✗ 오류: {e}\n", "err")
                results.append({"id": ch, "title": ch, "status": "error", "error": str(e)})

        self.app.channel_results = results
        win.finish(True)
        self.after(0, lambda: self.run_btn.configure(state="normal"))
        self.after(0, lambda: self._status.set(f"완료 ({len(results)}개)"))
        self.after(0, lambda: self.app.switch_tab("dashboard"))
        self.after(0, lambda: self.app.dashboard.show_channel())


# ── 영상 분석 탭 ──────────────────────────────────────────────────────────────
class VideoTab(tk.Frame):
    def __init__(self, master, app):
        super().__init__(master, bg=BG)
        self.app = app
        self._build()

    def _build(self):
        _section_header(self, "단일 영상 분석",
                         "YouTube URL / 영상 ID → 조회수 · 좋아요 · 댓글 수 수집")

        wrap = tk.Frame(self, bg=BG, padx=28)
        wrap.pack(fill="both", expand=True)

        tk.Label(wrap, text="영상 목록  (한 줄에 하나 · URL 또는 11자리 ID)",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(16, 4))
        border = tk.Frame(wrap, bg=ACCENT, padx=1, pady=1)
        border.pack(fill="x")
        self.vid_txt = tk.Text(border, height=8, font=("Consolas", 10),
                               bg=BG3, fg=FG, insertbackground=ACCENT,
                               relief="flat", padx=10, pady=8)
        self.vid_txt.pack(fill="both")

        btn_row = tk.Frame(wrap, bg=BG, pady=10)
        btn_row.pack(fill="x")
        self.run_btn = _btn(btn_row, "▶  분석 시작", self._run,
                            bg=ACCENT, fg="white", bold=True, padx=20, pady=10)
        self.run_btn.pack(side="left")

        self._status = tk.StringVar(value="대기 중")
        tk.Label(btn_row, textvariable=self._status, font=("Consolas", 9),
                 bg=BG, fg=FG_DIM).pack(side="left", padx=14)

    def _run(self):
        from youtube_api import extract_video_id
        lines = [
            ln.strip() for ln in self.vid_txt.get("1.0", "end").splitlines() if ln.strip()
        ]
        ids = list(dict.fromkeys(extract_video_id(l) for l in lines))
        ids = [i for i in ids if len(i) == 11]
        if not ids:
            messagebox.showwarning("입력 필요", "올바른 영상 URL 또는 ID를 입력하세요.")
            return
        if not _load_yt_api_key():
            messagebox.showerror("API Key 없음", "YouTube API Key가 설정되지 않았습니다.")
            return

        self.run_btn.configure(state="disabled")
        self._status.set("분석 중...")

        win = ConsoleWindow(self.winfo_toplevel(), "단일 영상 분석")
        win.append(f"[시작] 영상 {len(ids)}개 분석\n", "info")

        threading.Thread(target=self._analyze, args=(ids, win), daemon=True).start()

    def _analyze(self, ids, win):
        from youtube_api import fetch_videos_by_ids
        try:
            results = fetch_videos_by_ids(ids)
            win.append(f"\n✓ {len(results)}개 영상 수집 완료\n", "ok")
            for r in results:
                win.append(
                    f"  {r['title'][:50]}  |  "
                    f"조회수 {fmt_num(r['viewCount'])}  좋아요 {fmt_num(r['likeCount'])}  "
                    f"댓글 {fmt_num(r['commentCount'])}\n"
                )
        except Exception as e:
            win.append(f"\n✗ 오류: {e}\n", "err")
            results = []

        self.app.video_results = results
        win.finish(bool(results))
        self.after(0, lambda: self.run_btn.configure(state="normal"))
        self.after(0, lambda: self._status.set(f"완료 ({len(results)}개)"))
        self.after(0, lambda: self.app.switch_tab("dashboard"))
        self.after(0, lambda: self.app.dashboard.show_video())


# ── Instagram 스크래퍼 (모듈 레벨) ────────────────────────────────────────────────

def _ig_sleep(a: float = 1.3, b: float = 2.8) -> None:
    time.sleep(random.uniform(a, b))

def _ig_normalize_url(url: str) -> str:
    return url.split("?")[0].rstrip("/") + "/"

def _ig_parse_number(text) -> "int | None":
    if not text:
        return None
    compact = str(text).replace(",", "").replace(" ", "").upper()
    for unit, mul in [("억", 100_000_000), ("만", 10_000), ("천", 1_000)]:
        m = re.search(rf"(\d+(?:\.\d+)?){unit}", compact)
        if m:
            try:
                return int(float(m.group(1)) * mul)
            except ValueError:
                return None
    m = re.search(r"(\d+(?:\.\d+)?)([KMB])\b", compact)
    if m:
        mul_map = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
        try:
            return int(float(m.group(1)) * mul_map[m.group(2)])
        except ValueError:
            return None
    m = re.search(r"(\d+)", compact)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None
    return None

def _ig_get_chrome_ver():
    import subprocess, re as _re2
    cmds = [
        r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version',
        r'reg query "HKEY_LOCAL_MACHINE\Software\Google\Chrome\BLBeacon" /v version',
        r'reg query "HKEY_LOCAL_MACHINE\Software\WOW6432Node\Google\Chrome\BLBeacon" /v version',
    ]
    for cmd in cmds:
        try:
            out = subprocess.check_output(cmd, shell=True, text=True,
                                          encoding="utf-8", errors="ignore")
            m = _re2.search(r"(\d+)\.\d+\.\d+\.\d+", out)
            if m:
                return int(m.group(1))
        except Exception:
            continue
    return None

def _ig_create_driver(headless: bool = False):
    import undetected_chromedriver as uc
    opts = uc.ChromeOptions()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=ko-KR")
    chrome_major = _ig_get_chrome_ver()
    try:
        if chrome_major:
            driver = uc.Chrome(options=opts, version_main=chrome_major, headless=headless)
        else:
            driver = uc.Chrome(options=opts, headless=headless)
    except Exception:
        driver = uc.Chrome(options=opts, headless=headless)
    driver.implicitly_wait(3)
    return driver

def _ig_save_cookies(driver, path: str) -> None:
    with open(path, "wb") as f:
        pickle.dump(driver.get_cookies(), f)

def _ig_load_cookies(driver, path: str) -> bool:
    import os
    if not os.path.exists(path):
        return False
    try:
        driver.get("https://www.instagram.com/")
        _ig_sleep(2, 3)
        with open(path, "rb") as f:
            cookies = pickle.load(f)
        for cookie in cookies:
            c = cookie.copy()
            if "sameSite" in c and c["sameSite"] not in ("Strict", "Lax", "None"):
                c.pop("sameSite", None)
            try:
                driver.add_cookie(c)
            except Exception:
                continue
        driver.get("https://www.instagram.com/")
        _ig_sleep(3, 4)
        return True
    except Exception:
        return False

def _ig_is_logged_in(driver) -> bool:
    try:
        driver.get("https://www.instagram.com/")
        _ig_sleep(2, 3)
        return "accounts/login" not in driver.current_url.lower()
    except Exception:
        return False

def _ig_dismiss_popups(driver) -> None:
    """팝업 버튼이 있으면 한 번만 탐지·클릭 (단일 XPath OR로 불필요한 대기 제거)."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    _texts = ["나중에 하기", "나중에", "취소", "Not Now", "Cancel"]
    _xp = " | ".join(
        f"//button[normalize-space()='{t}'] | //div[@role='button' and normalize-space()='{t}']"
        for t in _texts
    )
    try:
        btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, _xp)))
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
        time.sleep(0.5)
    except Exception:
        pass

def _ig_login(driver, ig_id: str, ig_pw: str, cookie_path: str) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    driver.get("https://www.instagram.com/accounts/login/")
    # Instagram은 name 속성이 "username"/"password" 또는 "email"/"pass" 로 변경될 수 있음
    WebDriverWait(driver, 15).until(
        lambda d: d.find_elements(By.CSS_SELECTOR, "input[name='username'], input[name='email']")
    )
    user_el = (
        driver.find_elements(By.NAME, "username") or
        driver.find_elements(By.NAME, "email")
    )[0]
    pw_el = (
        driver.find_elements(By.NAME, "password") or
        driver.find_elements(By.NAME, "pass")
    )[0]
    user_el.clear(); user_el.send_keys(ig_id)
    _ig_sleep(0.5, 1.0)
    pw_el.clear(); pw_el.send_keys(ig_pw)
    _ig_sleep(0.5, 1.0)
    pw_el.send_keys(Keys.ENTER)
    time.sleep(6)
    _ig_dismiss_popups(driver)
    if not _ig_is_logged_in(driver):
        raise RuntimeError("Instagram 로그인 실패 — ID/PW 확인 또는 2FA 비활성화 필요")
    _ig_save_cookies(driver, cookie_path)

def _ig_ensure_login(driver, ig_id: str, ig_pw: str, cookie_path: str) -> None:
    if _ig_load_cookies(driver, cookie_path) and _ig_is_logged_in(driver):
        return
    _ig_login(driver, ig_id, ig_pw, cookie_path)

def _ig_open_reels_tab(driver, username: str) -> None:
    """릴스 탭으로 이동. /username/reels/ 직접 이동 → 실패 시 프로필에서 탭 클릭."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    reels_url = f"https://www.instagram.com/{username}/reels/"
    driver.get(reels_url)
    _ig_dismiss_popups(driver)
    # 릴스 링크가 1개라도 나타나면 바로 진행 (최대 6s 대기)
    try:
        WebDriverWait(driver, 6).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "a[href*='/reel/']")
        )
        if f"/{username.lower()}/reels" in driver.current_url.lower():
            return
    except Exception:
        pass

    # 백업: 프로필 진입 후 릴스 탭 클릭
    driver.get(f"https://www.instagram.com/{username}/")
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "header"))
    )
    css_selectors = [
        f"a[href='/{username}/reels/']",
        f"a[href='/{username.lower()}/reels/']",
        "a[href$='/reels/']",
    ]
    for sel in css_selectors:
        try:
            el = WebDriverWait(driver, 4).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
            )
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            WebDriverWait(driver, 6).until(
                lambda d: d.find_elements(By.CSS_SELECTOR, "a[href*='/reel/']")
            )
            return
        except Exception:
            pass
    # XPath 백업
    try:
        el = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((
            By.XPATH,
            "//a[contains(@href, '/reels/') and (@role='link' or @tabindex='0')]",
        )))
        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)
        WebDriverWait(driver, 6).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "a[href*='/reel/']")
        )
    except Exception:
        driver.get(reels_url)
        _ig_sleep(2, 3)


def _ig_collect_reel_links(driver, username: str, max_reels: int = 10) -> list:
    """/username/reels/ 탭으로 이동 후 /reel/ 링크만 수집."""
    from selenium.webdriver.common.by import By
    _ig_open_reels_tab(driver, username)
    links: dict = {}  # url → None, 삽입 순서(최신순) 유지 + 중복 제거
    retry = last_count = 0
    while len(links) < max_reels and retry < 6:
        for elem in driver.find_elements(By.CSS_SELECTOR, "a[href*='/reel/']"):
            href = elem.get_attribute("href")
            if href and "/reel/" in href:
                links[_ig_normalize_url(href)] = None
        if len(links) == last_count:
            retry += 1
        else:
            retry = 0
            last_count = len(links)
        driver.execute_script("window.scrollBy(0, 1400);")
        _ig_sleep(1.0, 1.6)
    return list(links)[:max_reels]

def _ig_scrape_reels_from_grid(driver, username: str, max_reels: int = 10) -> list:
    """릴스 탭 그리드에서 개별 페이지 방문 없이 지표 수집.
    - 조회수: 그리드 셀 내 span[dir] 안의 html-span (항상 노출)
    - 좋아요·댓글: 셀에 마우스오버 후 dir 없는 html-span 숫자들 (좋아요, 댓글 순)
    탐색 범위는 <a> 태그가 아닌 '그리드 셀 컨테이너' 기준으로 수행.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.action_chains import ActionChains

    _ig_open_reels_tab(driver, username)
    _ig_dismiss_popups(driver)

    results: dict = {}  # url -> row (삽입 순서 = 최신순)
    retry = last_count = 0
    actions = ActionChains(driver)

    def _cell_of(anchor):
        """anchor의 가장 가까운 그리드 셀(reel 링크가 하나만 있는 최소 조상) 반환."""
        return driver.execute_script(
            "let e = arguments[0].parentElement;"
            "while(e && e.querySelectorAll('a[href*=\"/reel/\"]').length !== 1)"
            "{ e = e.parentElement; }"
            "return e || arguments[0].parentElement;",
            anchor,
        )

    def _parse_nums(elements):
        """span 목록에서 숫자 파싱 (만/천/K/M 등 모두 처리)."""
        nums = []
        for s in elements:
            t = s.text.strip()
            if t:
                n = _ig_parse_number(t)
                if n is not None:
                    nums.append(n)
        return nums

    while len(results) < max_reels and retry < 6:
        anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/reel/']")
        for a in anchors:
            if len(results) >= max_reels:
                break
            href = a.get_attribute("href")
            if not href or "/reel/" not in href:
                continue
            url = _ig_normalize_url(href)
            if url in results:
                continue

            # 그리드 셀 컨테이너 (stats overlay가 <a> 바깥에 있을 수 있음)
            try:
                cell = _cell_of(a)
            except Exception:
                cell = a

            # 조회수 — dir="auto" wrapper 내부 html-span (항상 노출)
            view_count = None
            try:
                vspan = cell.find_element(
                    By.XPATH, ".//span[@dir]//span[contains(@class,'html-span')]"
                )
                view_count = _ig_parse_number(vspan.text.strip())
            except Exception:
                pass

            # 썸네일
            thumbnail_url = ""
            try:
                img = a.find_element(By.TAG_NAME, "img")
                thumbnail_url = img.get_attribute("src") or ""
            except Exception:
                pass

            # 마우스오버 → 좋아요·댓글 (dir 없는 html-span, 셀 전체 기준)
            like_count = comment_count = None
            try:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center',behavior:'instant'});", a
                )
                _ig_sleep(0.15, 0.3)
                actions.move_to_element(a).perform()
                _ig_sleep(0.6, 1.0)
                hover_spans = cell.find_elements(
                    By.XPATH,
                    ".//span[contains(@class,'html-span')"
                    " and not(ancestor::span[@dir])]",
                )
                nums = _parse_nums(hover_spans)
                if len(nums) >= 2:
                    like_count, comment_count = nums[0], nums[1]
                elif len(nums) == 1:
                    like_count = nums[0]
            except Exception:
                pass

            results[url] = {
                "account":        username,
                "reel_url":       url,
                "caption":        "",
                "thumbnail_url":  thumbnail_url,
                "video_url":      "",
                "like_count":     like_count,
                "like_raw":       str(like_count) if like_count is not None else "",
                "view_count":     view_count,
                "view_raw":       str(view_count) if view_count is not None else "",
                "comment_count":  comment_count,
                "comment_raw":    str(comment_count) if comment_count is not None else "",
                "posted_at":      "",
                "scraped_at":     datetime.now().isoformat(),
            }

        if len(results) == last_count:
            retry += 1
        else:
            retry = 0
            last_count = len(results)
        if len(results) < max_reels:
            driver.execute_script("window.scrollBy(0, 1400);")
            _ig_sleep(1.0, 1.6)

    return list(results.values())[:max_reels]


def _ig_scrape_post(driver, post_url: str, account: str) -> dict:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    driver.get(post_url)
    _ig_dismiss_popups(driver)

    # 날짜 — 페이지 핵심 요소 로드 대기 겸용
    posted_at = ""
    try:
        tel = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "time[datetime]"))
        )
        posted_at = (tel.get_attribute("datetime") or "")[:10]
    except Exception:
        _ig_sleep(1.0, 1.8)  # time 태그 없으면 최소 대기

    # 캡션
    caption = ""
    for sel in ["h1", "article h1", "div[role='dialog'] h1"]:
        try:
            caption = driver.find_element(By.CSS_SELECTOR, sel).text.strip()
            if caption:
                break
        except Exception:
            pass
    if not caption:
        try:
            caption = driver.find_element(By.TAG_NAME, "article").text.strip()[:500]
        except Exception:
            pass

    # 이미지 / 비디오
    all_image_urls: list = []
    seen: set = set()
    for img in driver.find_elements(By.CSS_SELECTOR, "article img, img"):
        src = img.get_attribute("src")
        if src and src.startswith("http") and src not in seen:
            seen.add(src)
            all_image_urls.append(src)
    video_url = ""
    for v in driver.find_elements(By.CSS_SELECTOR, "video"):
        src = v.get_attribute("src")
        if src and src.startswith("http"):
            video_url = src
            break

    # 좋아요 / 조회수 / 댓글수
    like_count = view_count = comment_count = None
    like_raw = view_raw = comment_raw = ""
    candidates = driver.find_elements(
        By.XPATH,
        "//*[contains(text(),'좋아요') or contains(text(),'likes') "
        "or contains(text(),'조회') or contains(text(),'views') "
        "or contains(text(),'댓글') or contains(text(),'comments')]",
    )
    for el in candidates:
        txt = (el.text or "").strip()
        if not txt:
            continue
        if like_count is None and ("좋아요" in txt or "like" in txt.lower()):
            like_raw   = txt
            like_count = _ig_parse_number(txt)
        if view_count is None and ("조회" in txt or "view" in txt.lower()):
            view_raw   = txt
            view_count = _ig_parse_number(txt)
        if comment_count is None and ("댓글" in txt or "comment" in txt.lower()):
            comment_raw   = txt
            comment_count = _ig_parse_number(txt)

    is_carousel = len(all_image_urls) >= 2

    return {
        "account":        account,
        "reel_url":       post_url,
        "caption":        caption[:200],
        "thumbnail_url":  all_image_urls[0] if all_image_urls else "",
        "video_url":      video_url,
        "like_count":     like_count,
        "like_raw":       like_raw,
        "view_count":     view_count,
        "view_raw":       view_raw,
        "comment_count":  comment_count,
        "comment_raw":    comment_raw,
        "posted_at":      posted_at,
        "scraped_at":     datetime.now().isoformat(),
    }


# ── 라이브 지표 크롤러 (모듈 레벨) ───────────────────────────────────────────────

def _parse_viewer_num(s: str) -> int:
    """'1,234' 또는 '1234' 형식의 문자열을 정수로 변환"""
    import re as _re
    s = _re.sub(r"[^\d]", "", str(s))
    return int(s) if s else 0


def _crawl_creator(platform: str, creator_id: str,
                   start_dt, end_dt,
                   categories: list, stop_event,
                   progress_cb=None) -> list:
    """
    viewership.softc.one에서 크리에이터의 방송 지표를 수집.
    날짜 쿼리 파라미터 URL + CSS 셀렉터 기반 SPA 파싱.

    platform  : 'chzzk' 또는 'soop'
    creator_id: 크리에이터 채널 ID
    """
    import re as _re
    import time as _time
    import random as _rnd
    import subprocess as _sp
    from urllib.parse import quote as _q
    from datetime import timedelta as _td
    from bs4 import BeautifulSoup as _BS
    import undetected_chromedriver as _uc
    from selenium.webdriver.common.by import By as _By
    from selenium.webdriver.support.ui import WebDriverWait as _Wait
    from selenium.webdriver.support import expected_conditions as _EC
    from selenium.common.exceptions import (
        TimeoutException as _TE,
        StaleElementReferenceException as _SRE,
    )

    def _log(msg):
        if progress_cb:
            progress_cb(msg)

    # ── Chrome 버전 감지 (reg query 방식 — winreg보다 Windows에서 안정적) ─────
    def _get_chrome_ver():
        cmds = [
            r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version',
            r'reg query "HKEY_LOCAL_MACHINE\Software\Google\Chrome\BLBeacon" /v version',
            r'reg query "HKEY_LOCAL_MACHINE\Software\WOW6432Node\Google\Chrome\BLBeacon" /v version',
        ]
        for cmd in cmds:
            try:
                out = _sp.check_output(cmd, shell=True, text=True, encoding="utf-8", errors="ignore")
                m = _re.search(r"(\d+)\.\d+\.\d+\.\d+", out)
                if m:
                    return int(m.group(1))
            except Exception:
                continue
        return None

    # ── URL (KST → UTC -9h, 날짜 쿼리 파라미터 포함) ─────────────────────────
    BASE      = "https://viewership.softc.one"
    PLAT_PATH = {"chzzk": "naverchzzk", "soop": "afreeca"}.get(platform, platform)
    start_utc = (start_dt.replace(hour=0,  minute=0,  second=0,  microsecond=0)    - _td(hours=9)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    end_utc   = (end_dt.replace(  hour=14, minute=59, second=59, microsecond=999000) - _td(hours=9)).strftime("%Y-%m-%dT%H:%M:%S.999Z")
    url = f"{BASE}/channel/{PLAT_PATH}/{creator_id}/streams?startDateTime={_q(start_utc)}&endDateTime={_q(end_utc)}"

    # ── CSS 셀렉터 ────────────────────────────────────────────────────────────
    STREAM_SEL   = ("a[href*='/streams/'] > button.min-h-11.py-2.hidden.lg\\:flex"
                    ".gap-4.text-xs.items-center.font-medium.leading-none"
                    ".rounded-lg.px-6.transition-all")
    PAGE_BTN_SEL = "button.font-inter.text-xs.w-8.h-8"

    # ── 드라이버 (버전 감지 성공 시 version_main 전달, 실패 시 UC 자체 감지) ──
    _log("    드라이버 시작 중...\n")
    chrome_major = _get_chrome_ver()
    _log(f"    감지된 Chrome major: {chrome_major if chrome_major else '자동감지'}\n")
    opts = _uc.ChromeOptions()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    try:
        if chrome_major:
            driver = _uc.Chrome(options=opts, version_main=chrome_major)
        else:
            driver = _uc.Chrome(options=opts)
        driver.implicitly_wait(3)
    except Exception as _e:
        import traceback as _tb
        _log(f"    [드라이버 오류] {_e}\n{_tb.format_exc()}\n")
        raise
    _log("    드라이버 준비 완료\n")

    # ── 페이지네이션 헬퍼 ─────────────────────────────────────────────────────
    def _num_page_btns():
        btns = driver.find_elements(_By.CSS_SELECTOR, PAGE_BTN_SEL)
        return [b for b in btns if (b.text or "").strip().isdigit()]

    def _click_page(target: str, timeout=5.0) -> bool:
        end_t = _time.time() + timeout
        while _time.time() < end_t:
            btns = _num_page_btns()
            btn  = next((b for b in btns if (b.text or "").strip() == target), None)
            if btn:
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    _time.sleep(0.2)
                    btn.click()
                    return True
                except (_SRE, Exception):
                    pass
            _time.sleep(0.2)
        return False

    def _wait_page_change(before_text: str, timeout=7.0) -> bool:
        end_t = _time.time() + timeout
        while _time.time() < end_t:
            try:
                elems = driver.find_elements(_By.CSS_SELECTOR, STREAM_SEL)
                after = (elems[0].text or "").strip() if elems else ""
                if after and after != before_text:
                    return True
            except Exception:
                pass
            _time.sleep(0.2)
        return False

    # ── 페이지 파싱 ───────────────────────────────────────────────────────────
    def _parse_page() -> list:
        soup = _BS(driver.page_source, "html.parser")
        rows = []
        for a in soup.select("a[href*='/streams/']"):
            btn = a.find("button")
            if not btn:
                continue
            cols = btn.find_all("div", recursive=False)
            if not cols:
                cols = btn.find_all("div")

            def _t(el):
                return el.get_text(strip=True) if el else ""

            def _n(el):
                s = _re.sub(r"[^\d]", "", _t(el))
                return int(s) if s else 0

            # 카테고리 / 제목
            col0 = cols[0] if cols else None
            divs = col0.find_all("div") if col0 else []
            cat_text   = _t(divs[0]) if len(divs) >= 1 else _t(col0)
            title_text = _t(divs[1]) if len(divs) >= 2 else ""

            # 기간 → 날짜 추출 (형식: "03.06 (금) 00:15 ~" — YYYY 없는 MM.DD)
            period   = _t(cols[1]) if len(cols) > 1 else ""
            date_m   = _re.search(r'(\d{1,2})\.(\d{2})', period)
            date_str = (f"{start_dt.year}-{int(date_m.group(1)):02d}-{int(date_m.group(2)):02d}"
                        if date_m else "")

            # 방송시간(h) → 분
            dur_text = _t(cols[2]) if len(cols) > 2 else ""
            dur_m    = _re.search(r'(\d+(?:\.\d+)?)', dur_text)
            dur_min  = int(float(dur_m.group(1)) * 60) if dur_m else 0

            peak = _n(cols[3]) if len(cols) > 3 else 0
            avg  = _n(cols[4]) if len(cols) > 4 else 0

            # 카테고리 필터
            if categories and cat_text and not any(c.lower() in cat_text.lower() for c in categories):
                continue

            rows.append({
                "creator":      creator_id,
                "platform":     platform.upper(),
                "title":        title_text,
                "category":     cat_text,
                "peak_viewers": peak,
                "avg_viewers":  avg,
                "date":         date_str,
                "duration_min": dur_min,
            })
        return rows

    # ── 메인 크롤링 루프 ──────────────────────────────────────────────────────
    results = []
    try:
        _log(f"    URL: {url}\n")
        driver.get(url)
        try:
            _Wait(driver, 15).until(
                _EC.presence_of_all_elements_located((_By.CSS_SELECTOR, STREAM_SEL))
            )
        except _TE:
            _log("    ⚠ 요소 대기 타임아웃 — 파싱 시도 계속\n")

        page = 1
        while not stop_event.is_set():
            _log(f"    {page}페이지 파싱 중...\n")
            rows = _parse_page()
            results.extend(rows)
            _log(f"    → {len(rows)}건\n")

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            _time.sleep(0.5)

            next_page = str(page + 1)
            if not any((b.text or "").strip() == next_page for b in _num_page_btns()):
                break

            before = ""
            try:
                elems  = driver.find_elements(_By.CSS_SELECTOR, STREAM_SEL)
                before = (elems[0].text or "").strip() if elems else ""
            except Exception:
                pass

            if not _click_page(next_page):
                break

            _wait_page_change(before)
            page += 1
            _time.sleep(_rnd.uniform(2.0, 4.0))

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    return results


# ── 라이브 지표 분석 탭 ────────────────────────────────────────────────────────
class LiveMetricsTab(tk.Frame):
    def __init__(self, master, app):
        super().__init__(master, bg=BG)
        self.app = app
        self._stop_event = threading.Event()
        self._thread = None
        self.live_results: list = []
        self._build()

    def _build(self):
        _section_header(self, "라이브 지표 분석",
                         "CHZZK / SOOP 방송 시청자 지표 수집  ·  viewership.softc.one")

        left = tk.Frame(self, bg=BG, padx=28)
        left.pack(fill="both", expand=True)

        # ── 크리에이터 ID 입력 ─────────────────────────────────────────────
        id_hdr = tk.Frame(left, bg=BG)
        id_hdr.pack(fill="x", pady=(16, 2))
        tk.Label(id_hdr, text="크리에이터 ID 목록  (한 줄에 하나)",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(side="left")
        _btn(id_hdr, "구글 시트에서 불러오기", self._load_from_gsheet,
             padx=10, pady=3).pack(side="right")
        tk.Label(left,
                 text="형식:  chzzk:채널ID  /  soop:아이디  /  URL 전체 붙여넣기 가능",
                 font=("Arial", 8), bg=BG, fg=FG_DIM, anchor="w").pack(fill="x", pady=(0, 4))

        id_border = tk.Frame(left, bg=ACCENT, padx=1, pady=1)
        id_border.pack(fill="x")
        self.id_txt = tk.Text(id_border, height=6, font=("Consolas", 10),
                              bg=BG3, fg=FG, insertbackground=ACCENT,
                              relief="flat", padx=10, pady=8)
        self.id_txt.pack(fill="both")

        # ── 날짜 범위 ──────────────────────────────────────────────────────
        date_row = tk.Frame(left, bg=BG, pady=8)
        date_row.pack(fill="x")

        tk.Label(date_row, text="시작일", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG).pack(side="left")
        self.start_date = tk.StringVar(
            value=(datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        tk.Entry(date_row, textvariable=self.start_date, width=12,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).pack(side="left", padx=(6, 16))

        tk.Label(date_row, text="종료일", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG).pack(side="left")
        self.end_date = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(date_row, textvariable=self.end_date, width=12,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).pack(side="left", padx=(6, 20))

        for lbl, days in [("7일", 7), ("30일", 30), ("90일", 90), ("전체", None)]:
            def _set_period(d=days):
                end = datetime.now()
                start = (end - timedelta(days=d)) if d else datetime(2020, 1, 1)
                self.start_date.set(start.strftime("%Y-%m-%d"))
                self.end_date.set(end.strftime("%Y-%m-%d"))
            _btn(date_row, lbl, _set_period, padx=10, pady=4).pack(side="left", padx=2)

        # ── 플랫폼 선택 ───────────────────────────────────────────────────
        opt_row = tk.Frame(left, bg=BG, pady=2)
        opt_row.pack(fill="x")

        tk.Label(opt_row, text="기본 플랫폼", font=("Arial", 9, "bold"),
                 bg=BG, fg=FG).pack(side="left", padx=(0, 8))
        self.platform_var = tk.StringVar(value="chzzk")
        for val, lbl in [("chzzk", "CHZZK"), ("soop", "SOOP")]:
            tk.Radiobutton(opt_row, text=lbl, variable=self.platform_var, value=val,
                           bg=BG, fg=FG_DIM, activebackground=BG,
                           selectcolor=BG3, font=("Arial", 9),
                           activeforeground=FG).pack(side="left", padx=4)

        self.headless_var = tk.BooleanVar(value=False)  # 항상 창 띄움

        # ── 버튼 행 ───────────────────────────────────────────────────────
        btn_row = tk.Frame(left, bg=BG, pady=6)
        btn_row.pack(fill="x")

        self.run_btn = _btn(btn_row, "▶  수집 시작", self._run,
                            bg=ACCENT, fg="white", bold=True, padx=20, pady=10)
        self.run_btn.pack(side="left")

        self.stop_btn = _btn(btn_row, "■  중지", self._stop, padx=14, pady=10)
        self.stop_btn.configure(state="disabled")
        self.stop_btn.pack(side="left", padx=6)

        self.export_btn = _btn(btn_row, "⬇  Excel 내보내기", self._export,
                               padx=14, pady=10)
        self.export_btn.pack(side="left", padx=6)

        self._status_var = tk.StringVar(value="대기 중")
        tk.Label(btn_row, textvariable=self._status_var,
                 font=("Consolas", 9), bg=BG, fg=FG_DIM).pack(side="left", padx=14)

        # ── 결과 Treeview ─────────────────────────────────────────────────
        tk.Frame(left, bg=BORDER, height=1).pack(fill="x", pady=(4, 0))
        tk.Label(left, text="수집 결과",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(6, 4))

        tree_frame = tk.Frame(left, bg=BG)
        tree_frame.pack(fill="both", expand=True)

        cols = ("플랫폼", "방송 제목", "카테고리",
                "최고 시청자", "평균 시청자", "날짜", "방송시간(분)")
        widths = (55, 220, 110, 80, 80, 85, 70)
        self.result_tree = ttk.Treeview(tree_frame, columns=cols,
                                        show="headings", style="Dark.Treeview", height=10)
        for i, col in enumerate(cols):
            self.result_tree.heading(col, text=col)
            self.result_tree.column(col, width=widths[i], minwidth=40, anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.result_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.result_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)


    # ── 구글 시트 불러오기 ─────────────────────────────────────────────────
    def _load_from_gsheet(self):
        """구글 스프레드시트에서 크리에이터 ID 목록을 불러와 textarea에 채운다.
        시트 열 구조: 크리에이터명 | 플랫폼 | URL
        공개(링크 공유) 시트에 한해 인증 없이 CSV export API를 사용한다.
        """
        import urllib.request
        import csv
        import io

        raw_url = _load_gsheet_url()
        if not raw_url:
            messagebox.showwarning("URL 없음",
                                   "구글 스프레드시트 URL이 설정되어 있지 않습니다.\n"
                                   "config.py 의 GSHEET_URL 에 URL을 설정하세요.")
            return

        # Sheet ID 추출
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", raw_url)
        if not m:
            messagebox.showerror("URL 오류", "올바른 구글 스프레드시트 URL이 아닙니다.")
            return
        sheet_id = m.group(1)

        # gid (탭 ID) 추출 – 없으면 첫 번째 시트(0)
        gid_m = re.search(r"gid=(\d+)", raw_url)
        gid = gid_m.group(1) if gid_m else "0"

        csv_url = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}"
            f"/export?format=csv&gid={gid}"
        )

        # CSV 다운로드
        try:
            req = urllib.request.Request(
                csv_url, headers={"User-Agent": "Mozilla/5.0"}
            )
            with urllib.request.urlopen(req, timeout=12) as resp:
                raw_csv = resp.read().decode("utf-8-sig")
        except Exception as e:
            messagebox.showerror(
                "불러오기 실패",
                f"스프레드시트를 가져올 수 없습니다.\n{e}\n\n"
                "시트가 '링크가 있는 모든 사용자'에게 공유되었는지 확인하세요.",
            )
            return

        reader = csv.DictReader(io.StringIO(raw_csv))
        rows = list(reader)
        fieldnames = reader.fieldnames or []

        if not rows:
            messagebox.showwarning("빈 시트", "시트에 데이터가 없습니다.")
            return

        # 열 이름 탐색 (대소문자·공백 무시)
        def _find_col(keywords):
            for f in fieldnames:
                if any(k in f.lower() for k in keywords):
                    return f
            return None

        plat_col = _find_col(["플랫폼", "platform"])
        url_col  = _find_col(["url"])

        if not url_col:
            messagebox.showerror(
                "열 없음",
                f"'URL' 열을 찾을 수 없습니다.\n발견된 열: {', '.join(fieldnames)}",
            )
            return

        default_plat = self.platform_var.get()
        lines = []

        for row in rows:
            raw = (row.get(url_col) or "").strip()
            if not raw:
                continue

            plat_raw = (row.get(plat_col) or "").strip().lower() if plat_col else ""

            # 플랫폼 판별 (열 값 우선, 없으면 URL로 추론)
            if "chzzk" in plat_raw or "chzzk" in raw:
                plat = "chzzk"
                m2 = re.search(
                    r"chzzk\.naver\.com/(?:live/)?([a-zA-Z0-9]+)", raw
                )
                cid = m2.group(1) if m2 else ""
            elif "soop" in plat_raw or "sooplive" in raw or "afreeca" in raw:
                plat = "soop"
                m2 = re.search(
                    r"(?:sooplive\.co\.kr|afreecatv\.com)/([^/?#\s]+)", raw
                )
                cid = m2.group(1) if m2 else ""
            else:
                # 플랫폼 불명: 기본 플랫폼으로 설정, URL 마지막 경로 세그먼트를 ID로
                plat = default_plat
                m2 = re.search(r"/([^/?#\s]+)/?$", raw.rstrip("/"))
                cid = m2.group(1) if m2 else ""

            if cid:
                lines.append(f"{plat}:{cid}")

        if not lines:
            messagebox.showwarning(
                "결과 없음",
                "URL 열에서 크리에이터 ID를 추출할 수 없었습니다.\n"
                "URL 형식을 확인하세요. (예: https://chzzk.naver.com/채널ID)",
            )
            return

        # ── 크리에이터별 그룹화 ───────────────────────────────────────────
        name_col = _find_col(["크리에이터", "이름", "name"])
        # rows와 lines는 빈 URL 행 스킵으로 인덱스가 불일치할 수 있으므로
        # rows를 직접 순회하면서 name과 (plat, cid) 쌍을 함께 추출
        from collections import OrderedDict
        grouped = OrderedDict()
        for row in rows:
            raw = (row.get(url_col) or "").strip()
            if not raw:
                continue
            plat_raw = (row.get(plat_col) or "").strip().lower() if plat_col else ""
            if "chzzk" in plat_raw or "chzzk" in raw:
                plat = "chzzk"
                m2 = re.search(r"chzzk\.naver\.com/(?:live/)?([a-zA-Z0-9]+)", raw)
                cid = m2.group(1) if m2 else ""
            elif "soop" in plat_raw or "sooplive" in raw or "afreeca" in raw:
                plat = "soop"
                m2 = re.search(r"(?:sooplive\.co\.kr|afreecatv\.com)/([^/?#\s]+)", raw)
                cid = m2.group(1) if m2 else ""
            else:
                plat = default_plat
                m2 = re.search(r"/([^/?#\s]+)/?$", raw.rstrip("/"))
                cid = m2.group(1) if m2 else ""
            if not cid:
                continue
            name = (row.get(name_col) or "").strip() if name_col else ""
            grouped.setdefault(name or cid, []).append((plat, cid))

        # 크리에이터명 기준 순서 유지 딕셔너리
        from collections import OrderedDict
        grouped: "OrderedDict[str, list]" = OrderedDict()
        for name, plat, cid in entries:
            grouped.setdefault(name, []).append((plat, cid))

        # ── 선택 다이얼로그 ───────────────────────────────────────────────
        selected = self._show_creator_select(grouped)
        if selected is None:   # 취소
            return

        result_lines = [f"{plat}:{cid}" for plat, cid in selected]
        if not result_lines:
            messagebox.showwarning("선택 없음", "선택된 크리에이터가 없습니다.")
            return

        self.id_txt.delete("1.0", "end")
        self.id_txt.insert("1.0", "\n".join(result_lines))

    # ── 크리에이터 선택 다이얼로그 ────────────────────────────────────────
    def _show_creator_select(self, grouped: dict):
        """
        grouped: {크리에이터명: [(plat, cid), ...]}
        반환: 선택된 [(plat, cid), ...] 또는 None(취소)
        """
        result = []
        cancelled = [False]

        win = tk.Toplevel(self)
        win.title("크리에이터 선택")
        win.configure(bg=BG)
        win.resizable(False, True)
        win.grab_set()

        # ── 헤더 ─────────────────────────────────────────────────────────
        tk.Label(win, text="조회할 크리에이터를 선택하세요",
                 font=("Arial", 10, "bold"), bg=BG, fg=FG,
                 padx=16, pady=10).pack(anchor="w")
        tk.Frame(win, bg=BORDER, height=1).pack(fill="x")

        # ── 스크롤 목록 ───────────────────────────────────────────────────
        list_frame = tk.Frame(win, bg=BG)
        list_frame.pack(fill="both", expand=True, padx=4, pady=4)

        canvas = tk.Canvas(list_frame, bg=BG, highlightthickness=0,
                           width=420, height=min(40 * len(grouped) + 8, 400))
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=BG)
        canvas_win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_win, width=e.width)
        inner.bind("<Configure>", _on_resize)

        # 체크박스 변수 (크리에이터명 단위)
        check_vars = {}  # {name: BooleanVar}
        rows_widgets = []

        PLAT_LABEL = {"chzzk": "CHZZK", "soop": "SOOP"}

        for name, pairs in grouped.items():
            var = tk.BooleanVar(value=True)
            check_vars[name] = var

            row_bg = BG3 if len(rows_widgets) % 2 == 0 else BG
            row = tk.Frame(inner, bg=row_bg, pady=3)
            row.pack(fill="x", padx=2, pady=1)

            tk.Checkbutton(row, variable=var, bg=row_bg, fg=FG,
                           activebackground=row_bg, selectcolor=BG3,
                           relief="flat").pack(side="left", padx=(6, 0))

            tk.Label(row, text=name, font=("Arial", 9, "bold"),
                     bg=row_bg, fg=FG, width=16, anchor="w").pack(side="left", padx=(2, 8))

            plat_str = "  ·  ".join(PLAT_LABEL.get(p, p.upper()) for p, _ in pairs)
            count_str = f"({len(pairs)})" if len(pairs) > 1 else ""
            tk.Label(row, text=plat_str, font=("Consolas", 9),
                     bg=row_bg, fg=ACCENT, anchor="w").pack(side="left")
            if count_str:
                tk.Label(row, text=f"  {count_str}", font=("Arial", 8),
                         bg=row_bg, fg=FG_DIM).pack(side="left")

            rows_widgets.append(row)

        # ── 하단 버튼 ─────────────────────────────────────────────────────
        tk.Frame(win, bg=BORDER, height=1).pack(fill="x")
        btn_row = tk.Frame(win, bg=BG, padx=12, pady=8)
        btn_row.pack(fill="x")

        def _all(v):
            for var in check_vars.values():
                var.set(v)

        _btn(btn_row, "전체 선택",  lambda: _all(True),  padx=10, pady=5).pack(side="left", padx=2)
        _btn(btn_row, "전체 해제",  lambda: _all(False), padx=10, pady=5).pack(side="left", padx=2)

        def _confirm():
            for name, var in check_vars.items():
                if var.get():
                    result.extend(grouped[name])
            win.destroy()

        def _cancel():
            cancelled[0] = True
            win.destroy()

        _btn(btn_row, "확인", _confirm,
             bg=ACCENT, fg="white", bold=True, padx=16, pady=5).pack(side="right", padx=2)
        _btn(btn_row, "취소", _cancel, padx=12, pady=5).pack(side="right", padx=2)

        win.bind("<Return>", lambda e: _confirm())
        win.bind("<Escape>", lambda e: _cancel())
        win.wait_window()

        return None if cancelled[0] else result

    # ── 입력 파싱 ──────────────────────────────────────────────────────────
    def _parse_ids(self) -> list:
        """
        textarea 내용을 (platform, creator_id) 튜플 리스트로 변환.
        지원 형식:
          chzzk:채널ID
          soop:아이디
          아이디          ← 기본 플랫폼 적용
        """
        default_plat = self.platform_var.get()
        lines = [
            ln.strip() for ln in self.id_txt.get("1.0", "end").splitlines()
            if ln.strip() and not ln.strip().startswith("#")
        ]
        result = []
        for line in lines:
            # viewership.softc.one URL 직접 붙여넣기 지원
            # 예: https://viewership.softc.one/channel/naverchzzk/ec857bee...
            if line.startswith("http"):
                import re as _re
                m = _re.search(r"softc\.one/channel/([^/]+)/([^/?#\s]+)", line)
                if m:
                    raw_plat, cid = m.group(1), m.group(2)
                    plat = "chzzk" if raw_plat == "naverchzzk" else raw_plat
                    if cid:
                        result.append((plat, cid))
                continue
            if ":" in line:
                plat, cid = line.split(":", 1)
                plat = plat.strip().lower()
                cid  = cid.strip().lstrip("@")
                if plat not in ("chzzk", "soop"):
                    plat = default_plat
            else:
                plat = default_plat
                cid  = line.lstrip("@").strip()
            if cid:
                result.append((plat, cid))
        return result

    # ── 수집 시작 ──────────────────────────────────────────────────────────
    def _run(self):
        creators = self._parse_ids()
        if not creators:
            messagebox.showwarning("입력 필요",
                                   "크리에이터 ID를 입력하세요.\n"
                                   "형식:  chzzk:채널ID  또는  soop:아이디")
            return

        try:
            start_dt = datetime.strptime(self.start_date.get(), "%Y-%m-%d")
            end_dt   = datetime.strptime(self.end_date.get(),   "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("날짜 오류", "날짜 형식: YYYY-MM-DD")
            return

        if start_dt > end_dt:
            messagebox.showerror("날짜 오류", "시작일이 종료일보다 늦습니다.")
            return

        # 초기화
        self.run_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal", bg=ACCENT, fg="white")
        self._status_var.set("수집 중...")
        self._stop_event.clear()
        self.live_results = []

        for row in self.result_tree.get_children():
            self.result_tree.delete(row)

        self._thread = threading.Thread(
            target=self._crawl_thread,
            args=(creators, start_dt, end_dt, []),
            daemon=True,
        )
        self._thread.start()

    # ── 중지 ──────────────────────────────────────────────────────────────
    def _stop(self):
        self._stop_event.set()
        self._done()

    def _done(self):
        self.run_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled", bg=BG3, fg=FG_DIM)

    # ── 크롤링 스레드 ──────────────────────────────────────────────────────
    def _crawl_thread(self, creators, start_dt, end_dt, categories):
        all_results = []
        total = len(creators)

        for idx, (platform, creator_id) in enumerate(creators, 1):
            if self._stop_event.is_set():
                break
            self.after(0, lambda i=idx, t=total: self._status_var.set(f"수집 중... ({i}/{t})"))
            try:
                rows = _crawl_creator(
                    platform, creator_id, start_dt, end_dt,
                    categories, self._stop_event, progress_cb=None,
                )
                all_results.extend(rows)

                for row in rows:
                    self.after(0, lambda r=row: self.result_tree.insert(
                        "", "end", values=(
                            r.get("platform",     ""),
                            r.get("title",        "")[:50],
                            r.get("category",     ""),
                            fmt_num(r.get("peak_viewers", 0)),
                            fmt_num(r.get("avg_viewers",  0)),
                            r.get("date",         ""),
                            r.get("duration_min", 0),
                        )
                    ))
            except Exception:
                pass

        self.live_results = all_results
        count = len(all_results)
        self.after(0, lambda: self._status_var.set(f"완료 ({count}건)"))
        self.after(0, self._done)
        # 대시보드 라이브 지표 탭 갱신
        self.after(0, lambda: self.app.dashboard.refresh_live())

    # ── Excel 내보내기 ──────────────────────────────────────────────────────
    def _export(self):
        if not self.live_results:
            messagebox.showwarning("데이터 없음", "수집된 데이터가 없습니다.\n먼저 수집을 실행하세요.")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"LiveMetrics_{ts}.xlsx",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("라이브러리 없음", "openpyxl 이 설치되어 있지 않습니다.\npip install openpyxl")
            return

        wb = openpyxl.Workbook()
        HDR_FILL = PatternFill("solid", fgColor="1A1A1A")
        HDR_FONT = Font(color="F4F4F5", bold=True)

        def _style_hdr(ws):
            for cell in ws[1]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="center")

        # ── 방송 목록 시트 ────────────────────────────────────────────────
        ws = wb.active
        ws.title = "방송 목록"
        ws.append(["플랫폼", "방송 제목", "카테고리",
                   "최고 시청자", "평균 시청자", "방송 날짜", "방송 시간(분)"])
        for r in self.live_results:
            ws.append([
                r.get("platform",     ""),
                r.get("title",        ""),
                r.get("category",     ""),
                r.get("peak_viewers", 0),
                r.get("avg_viewers",  0),
                r.get("date",         ""),
                r.get("duration_min", 0),
            ])
        _style_hdr(ws)

        # ── 크리에이터 요약 시트 ──────────────────────────────────────────
        ws2 = wb.create_sheet("크리에이터 요약")
        ws2.append(["크리에이터", "플랫폼", "방송 수",
                    "총 최고 시청자", "평균 최고 시청자",
                    "평균 시청자", "주요 카테고리"])

        from collections import defaultdict
        creator_data: dict = defaultdict(list)
        for r in self.live_results:
            key = (r.get("creator", ""), r.get("platform", ""))
            creator_data[key].append(r)

        for (creator, platform), rows in sorted(creator_data.items()):
            peaks = [r.get("peak_viewers", 0) for r in rows]
            avgs  = [r.get("avg_viewers",  0) for r in rows]
            cats  = [r.get("category", "") for r in rows if r.get("category")]
            top_cat = max(set(cats), key=cats.count) if cats else ""
            ws2.append([
                creator,
                platform,
                len(rows),
                sum(peaks),
                round(sum(peaks) / len(peaks)) if peaks else 0,
                round(sum(avgs)  / len(avgs))  if avgs  else 0,
                top_cat,
            ])
        _style_hdr(ws2)

        # ── 카테고리 집계 시트 ────────────────────────────────────────────
        ws3 = wb.create_sheet("카테고리 집계")
        ws3.append(["카테고리", "방송 수", "평균 최고 시청자", "평균 시청자"])

        cat_data: dict = defaultdict(list)
        for r in self.live_results:
            cat = r.get("category", "기타") or "기타"
            cat_data[cat].append(r)

        for cat, rows in sorted(cat_data.items(), key=lambda x: -len(x[1])):
            peaks = [r.get("peak_viewers", 0) for r in rows]
            avgs  = [r.get("avg_viewers",  0) for r in rows]
            ws3.append([
                cat,
                len(rows),
                round(sum(peaks) / len(peaks)) if peaks else 0,
                round(sum(avgs)  / len(avgs))  if avgs  else 0,
            ])
        _style_hdr(ws3)

        wb.save(path)
        messagebox.showinfo("저장 완료", f"Excel 저장 완료\n{path}")


# ── Instagram 분석 탭 ─────────────────────────────────────────────────────────
class InstagramTab(tk.Frame):
    def __init__(self, master, app):
        super().__init__(master, bg=BG)
        self.app = app
        self._stop_event = threading.Event()
        self._thread = None
        self._driver = None
        self.ig_results: list = []
        self._tree_row = [0]
        self._build()

    def _build(self):
        _section_header(self, "Instagram 릴스 분석",
                        "undetected-chromedriver  →  릴스 탭(/reels/)  →  지표 수집")

        wrap = tk.Frame(self, bg=BG, padx=28)
        wrap.pack(fill="both", expand=True)

        # ── 계정 목록 ────────────────────────────────────────────────────────
        tk.Label(wrap, text="Instagram 계정 목록  (한 줄에 하나, @ 제외)",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(16, 2))
        tk.Label(wrap, text="예:  nike  /  chanelofficial  /  hermes",
                 font=("Arial", 8), bg=BG, fg=FG_DIM, anchor="w").pack(fill="x", pady=(0, 4))
        acc_border = tk.Frame(wrap, bg=ACCENT, padx=1, pady=1)
        acc_border.pack(fill="x")
        self.acc_txt = tk.Text(acc_border, height=5, font=("Consolas", 10),
                               bg=BG3, fg=FG, insertbackground=ACCENT,
                               relief="flat", padx=10, pady=8)
        self.acc_txt.pack(fill="both")

        # ── 인증 정보 ────────────────────────────────────────────────────────
        cred_row = tk.Frame(wrap, bg=BG, pady=8)
        cred_row.pack(fill="x")
        tk.Label(cred_row, text="Instagram ID",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG).pack(side="left")
        self._ig_id = tk.StringVar()
        tk.Entry(cred_row, textvariable=self._ig_id, width=18,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).pack(side="left", padx=(6, 18))
        tk.Label(cred_row, text="Password",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG).pack(side="left")
        self._ig_pw = tk.StringVar()
        tk.Entry(cred_row, textvariable=self._ig_pw, width=18,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10), show="●").pack(side="left", padx=(6, 0))

        # ── 옵션 ─────────────────────────────────────────────────────────────
        opt_row = tk.Frame(wrap, bg=BG, pady=4)
        opt_row.pack(fill="x")
        tk.Label(opt_row, text="릴스 수 (계정당)",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG).pack(side="left")
        self._max_posts = tk.StringVar(value="10")
        tk.Entry(opt_row, textvariable=self._max_posts, width=5,
                 bg=BG3, fg=FG, insertbackground=ACCENT, relief="flat",
                 font=("Consolas", 10)).pack(side="left", padx=(6, 20))
        self._headless = tk.BooleanVar(value=False)
        tk.Checkbutton(opt_row, text="헤드리스  (Chrome 창 없이 실행)",
                       variable=self._headless, bg=BG, fg=FG_DIM,
                       activebackground=BG, selectcolor=BG3,
                       font=("Arial", 9), activeforeground=FG).pack(side="left")

        # ── 버튼 행 ──────────────────────────────────────────────────────────
        btn_row = tk.Frame(wrap, bg=BG, pady=6)
        btn_row.pack(fill="x")
        self.run_btn = _btn(btn_row, "▶  수집 시작", self._run,
                            bg=ACCENT, fg="white", bold=True, padx=20, pady=10)
        self.run_btn.pack(side="left")
        self.stop_btn = _btn(btn_row, "■  중지", self._stop, padx=14, pady=10)
        self.stop_btn.configure(state="disabled")
        self.stop_btn.pack(side="left", padx=6)
        self.export_btn = _btn(btn_row, "⬇  Excel 내보내기", self._export,
                               padx=14, pady=10)
        self.export_btn.pack(side="left", padx=6)
        self._status_var = tk.StringVar(value="대기 중")
        tk.Label(btn_row, textvariable=self._status_var,
                 font=("Consolas", 9), bg=BG, fg=FG_DIM).pack(side="left", padx=14)

        # ── 결과 Treeview ─────────────────────────────────────────────────────
        tk.Frame(wrap, bg=BORDER, height=1).pack(fill="x", pady=(8, 0))
        tk.Label(wrap, text="수집 결과",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x", pady=(6, 4))
        tree_frame = tk.Frame(wrap, bg=BG)
        tree_frame.pack(fill="both", expand=True)

        cols   = ("계정", "좋아요", "조회수", "댓글수", "게시일", "캡션")
        widths = (120,    80,      80,       70,       90,      290)
        self.result_tree = ttk.Treeview(tree_frame, columns=cols,
                                        show="headings", style="Dark.Treeview", height=10)
        for i, col in enumerate(cols):
            self.result_tree.heading(col, text=col)
            self.result_tree.column(col, width=widths[i], minwidth=40, anchor="w")
        self.result_tree.tag_configure("odd",  background=BG3)
        self.result_tree.tag_configure("even", background=BG5)
        _orig = self.result_tree.insert
        def _striped(*a, **kw):
            kw["tags"] = list(kw.get("tags", ())) + \
                         ["odd" if self._tree_row[0] % 2 == 0 else "even"]
            r = _orig(*a, **kw)
            self._tree_row[0] += 1
            return r
        self.result_tree.insert = _striped

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.result_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.result_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    # ── 수집 실행 ─────────────────────────────────────────────────────────────
    def _run(self):
        def _parse_account(raw: str) -> str:
            s = raw.strip().lstrip("@")
            # URL 형식이면 경로에서 첫 세그먼트 추출
            # e.g. https://www.instagram.com/eia_asmr/  →  eia_asmr
            m = re.search(r"instagram\.com/([^/?#]+)", s)
            if m:
                return m.group(1).strip("/")
            return s

        accounts = [
            _parse_account(ln)
            for ln in self.acc_txt.get("1.0", "end").splitlines()
            if ln.strip() and not ln.strip().startswith("#")
        ]
        if not accounts:
            messagebox.showwarning("입력 필요", "Instagram 계정을 입력하세요.")
            return
        ig_id = self._ig_id.get().strip()
        ig_pw = self._ig_pw.get().strip()
        if not ig_id or not ig_pw:
            messagebox.showwarning("인증 필요", "Instagram ID와 비밀번호를 입력하세요.")
            return
        try:
            max_posts = max(1, int(self._max_posts.get()))
        except ValueError:
            max_posts = 10

        self.run_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal", bg=ACCENT, fg="white")
        self._status_var.set("수집 중...")
        self._stop_event.clear()
        self.ig_results = []
        self._tree_row[0] = 0
        for row in self.result_tree.get_children():
            self.result_tree.delete(row)

        cookie_path = str(SCRIPT_DIR / "instagram_cookies.pkl")
        self._thread = threading.Thread(
            target=self._crawl_thread,
            args=(accounts, ig_id, ig_pw, max_posts, cookie_path),
            daemon=True,
        )
        self._thread.start()

    def _crawl_thread(self, accounts, ig_id, ig_pw, max_posts, cookie_path):
        try:
            self.after(0, lambda: self._status_var.set("드라이버 시작 중..."))
            driver = _ig_create_driver(headless=self._headless.get())
            self._driver = driver
            try:
                _ig_ensure_login(driver, ig_id, ig_pw, cookie_path)
                total = len(accounts)
                all_results = []
                for idx, account in enumerate(accounts, 1):
                    if self._stop_event.is_set():
                        break
                    self.after(0, lambda i=idx, t=total, a=account:
                               self._status_var.set(f"수집 중... ({i}/{t})  @{a}"))
                    try:
                        rows = _ig_scrape_reels_from_grid(driver, account, max_posts)
                    except Exception:
                        continue
                    for row in rows:
                        if self._stop_event.is_set():
                            break
                        all_results.append(row)
                        self.after(0, lambda r=row: self.result_tree.insert(
                            "", "end", values=(
                                r["account"],
                                fmt_num(r["like_count"])    if r["like_count"]    else "-",
                                fmt_num(r["view_count"])    if r["view_count"]    else "-",
                                fmt_num(r["comment_count"]) if r["comment_count"] else "-",
                                r["posted_at"],
                                r["caption"][:60],
                            )
                        ))
                    _ig_sleep(2.0, 3.5)
                self.ig_results = all_results
                self.app.instagram_results = all_results
            finally:
                try:
                    driver.quit()
                except Exception:
                    pass
                self._driver = None
            count = len(self.ig_results)
            self.after(0, lambda: self._status_var.set(f"완료 ({count}건)"))
            self.after(0, lambda: self.app.dashboard.show_instagram())
        except Exception as e:
            self.after(0, lambda: self._status_var.set(f"오류: {e}"))
        finally:
            self.after(0, self._done)

    def _stop(self):
        self._stop_event.set()
        if self._driver:
            try:
                self._driver.quit()
            except Exception:
                pass
            self._driver = None
        self._done()

    def _done(self):
        self.run_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled", bg=BG3, fg=FG_DIM)

    # ── Excel 내보내기 ─────────────────────────────────────────────────────────
    def _export(self):
        if not self.ig_results:
            messagebox.showwarning("데이터 없음", "수집된 데이터가 없습니다.")
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=f"Instagram_{ts}.xlsx",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("라이브러리 없음",
                                 "openpyxl 이 필요합니다.\npip install openpyxl")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instagram"
        HDR_FILL = PatternFill("solid", fgColor="1A1A1A")
        headers = [
            ("account","계정"), ("reel_url","릴스URL"),
            ("caption","캡션"), ("thumbnail_url","썸네일URL"), ("video_url","비디오URL"),
            ("like_count","좋아요"), ("view_count","조회수"), ("comment_count","댓글수"),
            ("posted_at","게시일"), ("scraped_at","수집일시"),
        ]
        for ci, (_, hdr) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = Font(bold=True, color="F4F4F5")
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(self.ig_results, 2):
            for ci, (key, _) in enumerate(headers, 1):
                ws.cell(row=ri, column=ci, value=row.get(key, ""))
        wb.save(path)
        messagebox.showinfo("저장 완료", f"저장되었습니다:\n{path}")


# ── 대시보드 탭 ───────────────────────────────────────────────────────────────
class DashboardTab(tk.Frame):
    def __init__(self, master, app):
        super().__init__(master, bg=BG)
        self.app = app
        self._current_sub = "channel"
        self._build()

    def _build(self):
        # ── 헤더 영역 ─────────────────────────────────────────────────────────
        top = tk.Frame(self, bg=BG2, padx=28, pady=16)
        top.pack(fill="x")
        hdr_left = tk.Frame(top, bg=BG2)
        hdr_left.pack(side="left", fill="y")
        tk.Frame(hdr_left, bg=ACCENT, width=3).pack(side="left", fill="y", padx=(0, 14))
        tk.Label(hdr_left, text="데이터 대시보드", font=("Arial", 15, "bold"),
                 bg=BG2, fg=FG).pack(side="left")
        self._excel_btn = _btn(top, "⬇  Excel 내보내기", self._export_excel,
                               bg=BG4, fg=FG_DIM, padx=14, pady=7)
        self._excel_btn.pack(side="right")
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # ── 서브탭 바 (언더라인 인디케이터) ──────────────────────────────────
        sub_bar = tk.Frame(self, bg=BG, padx=24, pady=0)
        sub_bar.pack(fill="x")
        self._sub_btns = {}
        for key, lbl in [("channel","채널 분석"), ("video","영상 분석"),
                         ("live","라이브 지표"), ("instagram","Instagram")]:
            tab_f = tk.Frame(sub_bar, bg=BG)
            tab_f.pack(side="left")
            btn = tk.Button(
                tab_f, text=lbl,
                command=lambda k=key: self._switch_sub(k),
                font=("Arial", 9), relief="flat", cursor="hand2",
                bg=BG, fg=FG_DIM,
                activebackground=BG, activeforeground=FG,
                padx=16, pady=10, bd=0,
            )
            btn.pack()
            ind = tk.Frame(tab_f, bg=BG, height=2)
            ind.pack(fill="x")
            self._sub_btns[key] = {"btn": btn, "ind": ind}
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # 컨텐츠 영역
        self._content = tk.Frame(self, bg=BG)
        self._content.pack(fill="both", expand=True)

        self._switch_sub("channel")

    def _switch_sub(self, key: str):
        self._current_sub = key
        for k, items in self._sub_btns.items():
            active = k == key
            items["btn"].configure(
                fg=FG if active else FG_DIM,
                font=("Arial", 9, "bold") if active else ("Arial", 9),
            )
            items["ind"].configure(bg=ACCENT if active else BG)
        for w in self._content.winfo_children():
            w.destroy()
        getattr(self, f"_build_{key}")()

    def show_channel(self):
        self._switch_sub("channel")

    def show_video(self):
        self._switch_sub("video")

    def show_instagram(self):
        self._switch_sub("instagram")

    # ── 채널 결과 ────────────────────────────────────────────────────────────
    def _build_channel(self):
        results = self.app.channel_results
        if not results:
            _label(self._content, "아직 채널 분석 결과가 없습니다.\n채널 통합 분석 탭에서 분석을 실행하세요.",
                   font_size=10, color=FG_DIM).pack(expand=True)
            return

        cols = ("채널명", "구독자", "쇼츠 avg", "쇼츠 수", "롱폼 avg", "롱폼 수", "상태")
        tree, _ = self._make_tree(cols, widths=(200,80,80,60,80,60,70))
        for r in results:
            tree.insert("", "end", values=(
                r.get("title", r.get("id","")),
                fmt_num(r.get("subscriberCount", 0)),
                fmt_num(r.get("avgShortsViews", 0)),
                r.get("shortsCount", 0),
                fmt_num(r.get("avgLongViews", 0)),
                r.get("longCount", 0),
                "완료" if r.get("status") == "completed" else f"오류: {r.get('error','')}",
            ), tags=(r.get("id",""),))

        def _on_select(e):
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            if idx < len(results):
                self._show_channel_detail(results[idx])
        tree.bind("<<TreeviewSelect>>", _on_select)

    def _show_channel_detail(self, r):
        if r.get("status") != "completed":
            return
        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"{r.get('title','')}  —  채널 상세")
        win.geometry("900x620")
        win.configure(bg=BG)

        # 헤더
        hdr = tk.Frame(win, bg=BG2, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text=r.get("title",""), font=("Arial", 14, "bold"),
                 bg=BG2, fg=FG).pack(side="left")
        tk.Label(hdr, text=f"구독자 {fmt_num(r.get('subscriberCount',0))}",
                 font=("Arial", 10), bg=BG2, fg=ACCENT).pack(side="left", padx=14)
        tk.Label(hdr, text=f"채널 ID: {r.get('id','')}",
                 font=("Consolas", 8), bg=BG2, fg=FG_DIM).pack(side="left")
        _btn(hdr, "YouTube에서 열기",
             lambda: webbrowser.open(f"https://youtube.com/channel/{r.get('id','')}"),
             padx=10, pady=6).pack(side="right")
        tk.Frame(win, bg=BORDER, height=1).pack(fill="x")

        nb = tk.Frame(win, bg=BG)
        nb.pack(fill="both", expand=True, padx=20, pady=10)

        # 쇼츠 / 롱폼 나란히
        left = tk.Frame(nb, bg=BG)
        left.pack(side="left", fill="both", expand=True, padx=(0,10))
        right = tk.Frame(nb, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        def _video_section(parent, title, videos, avg, is_short):
            tk.Label(parent, text=f"{title}  ({len(videos)}개)  avg {fmt_num(avg)}",
                     font=("Arial", 10, "bold"), bg=BG, fg=FG).pack(anchor="w", pady=(0,6))
            tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=(0,6))
            cols = ("제목", "조회수", "게시일")
            tree, _ = self._make_tree(cols, widths=(220,70,80), parent=parent, height=12)
            for v in videos:
                pub = v.get("publishedAt","")[:10]
                tree.insert("", "end", values=(
                    v.get("title","")[:50],
                    fmt_num(v.get("viewCount",0)),
                    pub,
                ), tags=(v.get("id",""),))
            def _open(e):
                sel = tree.selection()
                if not sel:
                    return
                idx = tree.index(sel[0])
                if idx < len(videos):
                    vid = videos[idx]
                    url = (f"https://youtube.com/shorts/{vid['id']}"
                           if vid.get("isShort") else f"https://youtu.be/{vid['id']}")
                    webbrowser.open(url)
            tree.bind("<Double-Button-1>", _open)

        _video_section(left, "쇼츠", r.get("shortsList",[]),
                       r.get("avgShortsViews",0), True)
        _video_section(right, "롱폼", r.get("longsList",[]),
                       r.get("avgLongViews",0), False)

    # ── 영상 결과 ────────────────────────────────────────────────────────────
    def _build_video(self):
        results = self.app.video_results
        if not results:
            _label(self._content, "아직 영상 분석 결과가 없습니다.\n단일 영상 분석 탭에서 분석을 실행하세요.",
                   font_size=10, color=FG_DIM).pack(expand=True)
            return

        cols = ("영상 제목", "채널", "유형", "조회수", "좋아요", "댓글", "게시일")
        tree, _ = self._make_tree(cols, widths=(250,120,50,80,70,70,80))
        for r in results:
            tree.insert("", "end", values=(
                r.get("title","")[:60],
                r.get("channelTitle","")[:30],
                "쇼츠" if r.get("isShort") else "롱폼",
                fmt_num(r.get("viewCount",0)),
                fmt_num(r.get("likeCount",0)),
                fmt_num(r.get("commentCount",0)),
                r.get("publishedAt","")[:10],
            ), tags=(r.get("videoId",""),))

        def _open(e):
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            if idx < len(results):
                v = results[idx]
                url = (f"https://youtube.com/shorts/{v['videoId']}"
                       if v.get("isShort") else f"https://youtube.com/watch?v={v['videoId']}")
                webbrowser.open(url)
        tree.bind("<Double-Button-1>", _open)

    # ── 라이브 지표 결과 ─────────────────────────────────────────────────────
    def _build_live(self):
        live_tab = self.app._pages.get("live")
        results = live_tab.live_results if live_tab else []
        if not results:
            _label(self._content,
                   "아직 라이브 지표 수집 결과가 없습니다.\n라이브 지표 분석 탭에서 수집을 실행하세요.",
                   font_size=10, color=FG_DIM).pack(expand=True)
            return

        # ── 플랫폼별 집계 ─────────────────────────────────────────────────
        from collections import defaultdict
        plat_rows: dict = defaultdict(list)
        for r in results:
            plat_rows[r.get("platform", "?").upper()].append(r)

        cols = ("플랫폼", "방송 수", "총 최고 시청자", "평균 최고 시청자", "평균 시청자")
        tree, _ = self._make_tree(cols, widths=(80, 70, 120, 120, 100), height=8)
        for plat, rows in sorted(plat_rows.items()):
            peaks = [r.get("peak_viewers", 0) for r in rows]
            avgs  = [r.get("avg_viewers",  0) for r in rows]
            tree.insert("", "end", values=(
                plat,
                len(rows),
                fmt_num(sum(peaks)),
                fmt_num(round(sum(peaks) / len(peaks))) if peaks else 0,
                fmt_num(round(sum(avgs)  / len(avgs)))  if avgs  else 0,
            ))

    # ── Instagram 결과 ───────────────────────────────────────────────────────
    def _build_instagram(self):
        results = self.app.instagram_results
        if not results:
            _label(self._content, "아직 Instagram 수집 결과가 없습니다.\nInstagram 분석 탭에서 수집을 실행하세요.",
                   font_size=10, color=FG_DIM).pack(expand=True)
            return

        # 계정별 집계
        from collections import defaultdict
        agg: dict = defaultdict(lambda: {"count": 0, "likes": [], "views": [], "comments": []})
        for r in results:
            acc = r.get("account", "")
            agg[acc]["count"] += 1
            if r.get("like_count") is not None:
                agg[acc]["likes"].append(r["like_count"])
            if r.get("view_count") is not None:
                agg[acc]["views"].append(r["view_count"])
            if r.get("comment_count") is not None:
                agg[acc]["comments"].append(r["comment_count"])

        summary_cols   = ("계정", "릴스수", "총 좋아요", "평균 좋아요", "총 조회수", "평균 조회수", "평균 댓글수")
        summary_widths = (130, 60, 90, 90, 90, 90, 90)
        tree, _ = self._make_tree(summary_cols, widths=summary_widths)

        for acc, d in agg.items():
            total_likes  = sum(d["likes"])
            avg_likes    = round(total_likes / len(d["likes"]))         if d["likes"]    else 0
            total_views  = sum(d["views"])
            avg_views    = round(total_views / len(d["views"]))         if d["views"]    else 0
            avg_comments = round(sum(d["comments"]) / len(d["comments"])) if d["comments"] else 0
            tree.insert("", "end", values=(
                acc, d["count"],
                fmt_num(total_likes)  if d["likes"]    else "-",
                fmt_num(avg_likes)    if d["likes"]    else "-",
                fmt_num(total_views)  if d["views"]    else "-",
                fmt_num(avg_views)    if d["views"]    else "-",
                fmt_num(avg_comments) if d["comments"] else "-",
            ))

        # 릴스 상세 테이블
        detail_lbl = tk.Frame(self._content, bg=BG, padx=20, pady=(8, 4))
        detail_lbl.pack(fill="x")
        tk.Label(detail_lbl, text="릴스 상세",
                 font=("Arial", 9, "bold"), bg=BG, fg=FG, anchor="w").pack(fill="x")

        detail_cols   = ("계정", "좋아요", "조회수", "댓글수", "게시일", "캡션")
        detail_widths = (120, 80, 80, 70, 90, 290)
        dtree, _ = self._make_tree(detail_cols, widths=detail_widths)
        for r in results:
            dtree.insert("", "end", values=(
                r.get("account", ""),
                fmt_num(r["like_count"])    if r.get("like_count")    is not None else "-",
                fmt_num(r["view_count"])    if r.get("view_count")    is not None else "-",
                fmt_num(r["comment_count"]) if r.get("comment_count") is not None else "-",
                r.get("posted_at", ""),
                r.get("caption", "")[:60],
            ))

        def _on_select(e):
            sel = dtree.selection()
            if not sel:
                return
            idx = dtree.index(sel[0])
            if idx < len(results):
                webbrowser.open(results[idx].get("reel_url", ""))

        dtree.bind("<<TreeviewSelect>>", _on_select)

    def refresh_live(self):
        """크롤링 완료 후 라이브 지표 서브탭 갱신 (현재 해당 탭이 열려있을 때만)"""
        if self._current_sub == "live":
            self._switch_sub("live")

    # ── Treeview 생성 헬퍼 ───────────────────────────────────────────────────
    def _make_tree(self, cols, widths=None, parent=None, height=16):
        parent = parent or self._content
        frame = tk.Frame(parent, bg=BG)
        frame.pack(fill="both", expand=True, padx=20 if parent is self._content else 0,
                   pady=(10, 0))

        tree = ttk.Treeview(frame, columns=cols, show="headings",
                            style="Dark.Treeview", height=height)
        for i, col in enumerate(cols):
            w = widths[i] if widths and i < len(widths) else 100
            tree.heading(col, text=col)
            tree.column(col, width=w, minwidth=40, anchor="w")

        # 줄무늬 행 (홀/짝 배경색 구분)
        tree.tag_configure("odd",  background=BG3)
        tree.tag_configure("even", background=BG5)
        _row = [0]
        _real_insert = tree.insert
        def _striped(*args, **kw):
            kw["tags"] = list(kw.get("tags", ())) + ["odd" if _row[0] % 2 == 0 else "even"]
            r = _real_insert(*args, **kw)
            _row[0] += 1
            return r
        tree.insert = _striped

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        return tree, frame

    # ── Excel 내보내기 ───────────────────────────────────────────────────────
    def _export_excel(self):
        sub = self._current_sub
        if sub == "channel" and not self.app.channel_results:
            messagebox.showwarning("데이터 없음", "채널 분석 결과가 없습니다.")
            return
        if sub == "video" and not self.app.video_results:
            messagebox.showwarning("데이터 없음", "영상 분석 결과가 없습니다.")
            return
        if sub == "live":
            live_tab = self.app._pages.get("live")
            if not live_tab or not live_tab.live_results:
                messagebox.showwarning("데이터 없음", "라이브 지표 수집 결과가 없습니다.")
                return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = {
            "channel": f"TubeMetric_Channel_{ts}.xlsx",
            "video":   f"TubeMetric_Video_{ts}.xlsx",
            "live":    f"TubeMetric_Live_{ts}.xlsx",
        }.get(sub, f"TubeMetric_{ts}.xlsx")

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("라이브러리 없음", "openpyxl 이 설치되어 있지 않습니다.")
            return

        wb = openpyxl.Workbook()

        HDR_FILL = PatternFill("solid", fgColor="1A1A1A")
        HDR_FONT = Font(color="F4F4F5", bold=True)

        def _style_header(ws):
            for cell in ws[1]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="center")

        if sub == "channel":
            ws = wb.active
            ws.title = "채널 요약"
            headers = ["채널 링크", "채널명", "채널 ID", "구독자 수",
                       "쇼츠 평균 조회수", "쇼츠 수", "롱폼 평균 조회수", "롱폼 수", "상태"]
            ws.append(headers)
            for r in self.app.channel_results:
                ws.append([
                    f"https://youtube.com/channel/{r.get('id','')}",
                    r.get("title",""),
                    r.get("id",""),
                    int(r.get("subscriberCount",0) or 0),
                    r.get("avgShortsViews",0),
                    r.get("shortsCount",0),
                    r.get("avgLongViews",0),
                    r.get("longCount",0),
                    "완료" if r.get("status") == "completed" else f"오류",
                ])
            _style_header(ws)

            for r in self.app.channel_results:
                if r.get("status") != "completed":
                    continue
                videos = r.get("shortsList",[]) + r.get("longsList",[])
                if not videos:
                    continue
                name = (r.get("title","")[:28]).replace("/","_")
                ws2 = wb.create_sheet(title=name or r.get("id","")[:28])
                ws2.append(["영상 링크","영상 제목","유형","조회수","게시일"])
                for v in videos:
                    url = (f"https://youtube.com/shorts/{v['id']}"
                           if v.get("isShort") else f"https://youtu.be/{v['id']}")
                    ws2.append([url, v.get("title",""), "쇼츠" if v.get("isShort") else "롱폼",
                                v.get("viewCount",0), v.get("publishedAt","")[:10]])
                _style_header(ws2)

        elif sub == "video":
            ws = wb.active
            ws.title = "영상 데이터"
            headers = ["영상 링크","영상 제목","채널명","유형","조회수","좋아요","댓글","게시일"]
            ws.append(headers)
            for r in self.app.video_results:
                url = (f"https://youtube.com/shorts/{r['videoId']}"
                       if r.get("isShort") else f"https://youtube.com/watch?v={r['videoId']}")
                ws.append([url, r.get("title",""), r.get("channelTitle",""),
                           "쇼츠" if r.get("isShort") else "롱폼",
                           r.get("viewCount",0), r.get("likeCount",0),
                           r.get("commentCount",0), r.get("publishedAt","")[:10]])
            _style_header(ws)

        elif sub == "live":
            from collections import defaultdict
            live_tab = self.app._pages.get("live")
            results = live_tab.live_results if live_tab else []
            # 집계 시트
            ws = wb.active
            ws.title = "플랫폼 집계"
            ws.append(["플랫폼","방송 수","총 최고 시청자","평균 최고 시청자","평균 시청자"])
            plat_rows: dict = defaultdict(list)
            for r in results:
                plat_rows[r.get("platform","?").upper()].append(r)
            for plat, rows in sorted(plat_rows.items()):
                peaks = [r.get("peak_viewers",0) for r in rows]
                avgs  = [r.get("avg_viewers", 0) for r in rows]
                ws.append([
                    plat, len(rows), sum(peaks),
                    round(sum(peaks)/len(peaks)) if peaks else 0,
                    round(sum(avgs) /len(avgs))  if avgs  else 0,
                ])
            _style_header(ws)
            # 전체 방송 목록 시트
            ws2 = wb.create_sheet("방송 목록")
            ws2.append(["플랫폼","방송 제목","카테고리","최고 시청자","평균 시청자","방송 날짜","방송 시간(분)"])
            for r in results:
                ws2.append([
                    r.get("platform",""), r.get("title",""), r.get("category",""),
                    r.get("peak_viewers",0), r.get("avg_viewers",0),
                    r.get("date",""), r.get("duration_min",0),
                ])
            _style_header(ws2)

        wb.save(path)
        messagebox.showinfo("저장 완료", f"Excel 파일이 저장되었습니다.\n{path}")


# ── 메인 앱 ──────────────────────────────────────────────────────────────────
class MainApp(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg=BG)
        self.channel_results = []
        self.video_results = []
        self.instagram_results = []
        self._build_ui()

    def _build_ui(self):
        _apply_treeview_style()

        # 사이드바
        sidebar = tk.Frame(self, bg=BG2, width=240)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # ── 로고 ────────────────────────────────────────────────────────────
        logo = tk.Frame(sidebar, bg=BG2, padx=20, pady=22)
        logo.pack(fill="x")
        logo_row = tk.Frame(logo, bg=BG2)
        logo_row.pack(anchor="w")
        tk.Frame(logo_row, bg=ACCENT, width=4, height=22).pack(side="left")
        tk.Label(logo_row, text="  TubeMetric",
                 font=("Arial", 13, "bold"), bg=BG2, fg=FG).pack(side="left")
        tk.Label(logo, text=f"@{MACHINE_INFO['operator']}",
                 font=("Arial", 8), bg=BG2, fg=FG_MUTE, anchor="w").pack(fill="x", pady=(6, 0))

        tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x")
        tk.Label(sidebar, text="NAVIGATION",
                 font=("Arial", 7, "bold"), bg=BG2, fg=FG_MUTE, anchor="w").pack(
                 fill="x", padx=22, pady=(12, 4))

        # ── 탭 버튼 (인디케이터 바 포함) ────────────────────────────────────
        self._tab_items: dict = {}
        self._tab_btns:  dict = {}
        for key, icon, label in [
            ("channel",   "▸", "채널 통합 분석"),
            ("video",     "▸", "단일 영상 분석"),
            ("live",      "▸", "라이브 지표 분석"),
            ("instagram", "▸", "Instagram 분석"),
            ("dashboard", "▸", "데이터 대시보드"),
        ]:
            row = tk.Frame(sidebar, bg=BG2)
            row.pack(fill="x")
            ind = tk.Frame(row, bg=BG2, width=3)
            ind.pack(side="left", fill="y")
            ind.pack_propagate(False)
            btn = tk.Button(
                row, text=f"  {icon}  {label}",
                command=lambda k=key: self.switch_tab(k),
                bg=BG2, fg=FG_DIM,
                activebackground=BG3, activeforeground=FG,
                font=("Arial", 10), relief="flat",
                anchor="w", padx=16, pady=12, cursor="hand2",
            )
            btn.pack(fill="both", expand=True)
            self._tab_items[key] = {"btn": btn, "ind": ind}
            self._tab_btns[key] = btn

        # ── 하단 시스템 정보 ─────────────────────────────────────────────────
        mf = tk.Frame(sidebar, bg=BG2)
        mf.pack(side="bottom", fill="x", padx=20, pady=14)
        tk.Frame(mf, bg=BORDER, height=1).pack(fill="x", pady=(0, 8))
        tk.Label(mf, text=MACHINE_INFO["hostname"],
                 font=("Arial", 8), bg=BG2, fg=FG_DIM, anchor="w").pack(fill="x")
        tk.Label(mf, text=f"{MACHINE_INFO['os'].split()[0]}  ·  v1.0",
                 font=("Arial", 7), bg=BG2, fg=FG_MUTE, anchor="w").pack(fill="x")

        # 컨텐츠
        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)

        self.dashboard = DashboardTab(self.content, self)

        self._pages = {
            "channel":   ChannelTab(self.content, self),
            "video":     VideoTab(self.content, self),
            "live":      LiveMetricsTab(self.content, self),
            "instagram": InstagramTab(self.content, self),
            "dashboard": self.dashboard,
        }
        self.switch_tab("channel")

    def switch_tab(self, key: str):
        for k, items in self._tab_items.items():
            active = k == key
            items["btn"].configure(
                bg=BG3 if active else BG2,
                fg=FG  if active else FG_DIM,
                font=("Arial", 10, "bold") if active else ("Arial", 10),
            )
            items["ind"].configure(bg=ACCENT if active else BG2)
        for k, page in self._pages.items():
            if k == key:
                page.pack(fill="both", expand=True)
            else:
                page.pack_forget()


# ── App ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PARABLE-TUBEMETRIC  |  YouTube Analytics")
        self.geometry("1100x680")
        self.minsize(900, 560)
        self.configure(bg=BG)
        self._show_pin()

    def _show_pin(self):
        for w in self.winfo_children():
            w.destroy()
        PinScreen(self, on_success=self._show_main).pack(fill="both", expand=True)

    def _show_main(self):
        for w in self.winfo_children():
            w.destroy()
        try:
            MainApp(self).pack(fill="both", expand=True)
        except Exception:
            import traceback
            err = traceback.format_exc()
            messagebox.showerror("초기화 오류",
                                 f"앱을 로드하는 중 오류가 발생했습니다.\n\n{err}")
            raise


if __name__ == "__main__":
    App().mainloop()
